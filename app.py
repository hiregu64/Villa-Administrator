import streamlit as st
import pandas as pd
import io
import datetime
import openpyxl
import json
import re
from openpyxl.styles import Font, Alignment
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

# ==============================================================================
# 1. STRUKTUREN & PARAMETER
# ==============================================================================
class KiAntwortSchema(BaseModel):
    wissensluecke_erkannt: bool = Field(description="True bei unvollständigem Excel-Kontext.")
    antwort_text: str = Field(description="Antworttext für Gast. Leerstring bei Wissenslücke.")

FILE_ID = '1FzhWZuO6aRZkdRuQBzaojhkq7bQDyprl'
FALLBACK_SATZ = "Ich habe dazu leider keine Informationen, Ich gebe das aber gern an die Hosts weiter."

st.set_page_config(page_title="Villa Avatar", page_icon="☀️", layout="centered")

# CSS Styling für Chat und Buttons
st.markdown("""
    <style>
    div.stButton > button[kind="primary"] { background-color: #e3f2fd !important; color: #1565c0 !important; border: 1px solid #bbdefb !important; font-weight: bold !important; }
    div[data-testid="stChatMessage"]:has(div[aria-label="Chat message from user"]) { flex-direction: row-reverse !important; background-color: #F0F2F6 !important; border-radius: 10px !important; padding: 10px !important; }
    div[data-testid="stChatMessage"]:has(div[aria-label="Chat message from user"]) div[data-testid="stChatMessageContent"] { text-align: right !important; width: 100% !important; }
    </style>
""", unsafe_allow_html=True)

# Session States flach initialisieren
for key, value in [
    ("aktive_rolle", None), ("aktiver_use_case", None), ("selected_object", None), 
    ("selected_field", None), ("messages", []), ("host_authentifiziert", False), 
    ("debug_modus_aktiv", False), ("last_write_status", "Noch kein Schreibvorgang."), 
    ("last_extracted_context", "Kein Kontext extrahiert."), ("matrix_data", None),
    ("erfolgsmeldung_anzeigen", None), ("host_text_wert", ""),
    ("selected_report_type", None), ("selected_report_timeframe", None)
]:
    if key not in st.session_state:
        st.session_state[key] = value

# ==============================================================================
# 2. DATEN-LADE ENGINE
# ==============================================================================
def fetch_matrix_from_drive():
    try:
        if "GOOGLE_CREDENTIALS" not in st.secrets:
            st.error("❌ Kritischer Fehler: 'GOOGLE_CREDENTIALS' fehlt in Secrets.")
            return False
            
        creds = service_account.Credentials.from_service_account_info(st.secrets["GOOGLE_CREDENTIALS"])
        service = build('drive', 'v3', credentials=creds)
        
        fh = io.BytesIO()
        MediaIoBaseDownload(fh, service.files().get_media(fileId=FILE_ID)).next_chunk()
        fh.seek(0)
        
        df_wissen = pd.read_excel(fh, sheet_name="Wissensbasis", header=0)
        fh.seek(0)
        df_lexikon = pd.read_excel(fh, sheet_name="Spalten_Lexikon", header=0)
        fh.seek(0)
        df_usecases = pd.read_excel(fh, sheet_name="UseCase_Lexikon", header=0)
        fh.seek(0)
        
        try: df_passwoerter = pd.read_excel(fh, sheet_name="Passwort_Lexikon", header=0)
        except: df_passwoerter = None
        
        if df_wissen is not None and "Wo?" in df_wissen.columns: 
            df_wissen["Wo?"] = df_wissen["Wo?"].ffill()
            
        for df in [df_usecases, df_passwoerter, df_wissen, df_lexikon]:
            if df is not None: 
                df.columns = [str(c).strip() for c in df.columns]
            
        st.session_state.matrix_data = {
            "wissen": df_wissen, "lexikon": df_lexikon, 
            "usecases": df_usecases, "passwoerter": df_passwoerter, "service": service
        }
        return True
    except Exception as e:
        st.error(f"❌ Ladefehler: {e}")
        return False

if st.session_state.matrix_data is None:
    with st.spinner("Daten werden geladen..."):
        if fetch_matrix_from_drive(): st.rerun()
        else: st.stop()

df_wissen = st.session_state.matrix_data["wissen"]
df_lexikon = st.session_state.matrix_data["lexikon"]
df_usecases = st.session_state.matrix_data["usecases"]
df_passwoerter = st.session_state.matrix_data["passwoerter"]
drive_service = st.session_state.matrix_data["service"]

def find_column_by_fuzzy_name(headers, target_name):
    cleaned = [str(h).strip().lower() for h in headers]
    search = str(target_name).strip().lower()
    return cleaned.index(search) + 1 if search in cleaned else None

# ==============================================================================
# 3. CHRONOLOGISCHE STATUS-PARSING ENGINE (Tolerant & Exakt)
# ==============================================================================
def parse_status_history(status_val):
    if pd.isna(status_val) or str(status_val).lower() == 'nan': return []
    
    raw_str = str(status_val).replace('\n', ' ')
    pattern = r'(\d{1,2}\.\d{1,2}\.\d{4})(?:\s+\d{2}:\d{2})?.*?\b(offen|ok|behoben|erfolgt|geschlossen|aktiv)\b'
    matches = re.findall(pattern, raw_str, re.IGNORECASE)
    
    parsed = []
    for dat_str, zustand_str in matches:
        try:
            parts = dat_str.split('.')
            d = int(parts[0])
            m = int(parts[1])
            y = int(parts[2])
            dt = datetime.datetime(y, m, d)
            
            z_clean = zustand_str.strip().lower()
            if z_clean in ["ok", "behoben", "erfolgt", "geschlossen"]:
                z_clean = "ok"
                
            parsed.append({"datum": dt, "zustand": z_clean})
        except:
            continue
            
    parsed.sort(key=lambda x: x["datum"])
    return parsed

def get_effective_days_excluding_winter(start_date, end_date):
    if start_date >= end_date: 
        return 0
    total_days = (end_date - start_date).days
    if total_days > 3650:
        return total_days
        
    winter_days = 0
    curr = start_date
    for _ in range(total_days + 1):
        if curr > end_date:
            break
        if curr.month in [11, 12, 1, 2, 3]:
            winter_days += 1
        curr += datetime.timedelta(days=1)
        
    return max(0, total_days - winter_days)

def parse_period_from_text(text_content):
    """
    Übersetzt Freitext-Intervalle intelligent in eine präzise Tagesanzahl.
    Berücksichtigt auch mehrere genannte Intervalle und wählt das kürzeste aus.
    """
    if pd.isna(text_content) or str(text_content).strip().lower() == "nan":
        return None
        
    text_clean = str(text_content).lower()
    gefundene_intervalle = []
    
    # 1. Textbasierte Schlagworte prüfen
    if "wöchentlich" in text_clean or "wochentlich" in text_clean:
        gefundene_intervalle.append(7)
    if "zweiwöchentlich" in text_clean or "zweiwochentlich" in text_clean:
        gefundene_intervalle.append(14)
    if "monatlich" in text_clean:
        gefundene_intervalle.append(30)
    if "vierteljährlich" in text_clean or "vierteljahrlich" in text_clean or "quartal" in text_clean:
        gefundene_intervalle.append(90)
    if "halbjährlich" in text_clean or "halbjahrlich" in text_clean:
        gefundene_intervalle.append(180)
    if "jährlich" in text_clean or "jahrlich" in text_clean:
        gefundene_intervalle.append(365)
        
    # 2. Zusätzlich nach reinen Zahlen suchen (z.B. 'alle 14 Tage')
    nums = re.findall(r'\d+', text_clean)
    for n in nums:
        val = int(n)
        if val > 0:
            gefundene_intervalle.append(val)
            
    return min(gefundene_intervalle) if gefundene_intervalle else None

# ==============================================================================
# 4. LLM KI-ENGINE
# ==============================================================================
def call_gemini(prompt, context="", structured=True):
    client = genai.Client(api_key=st.secrets.get("GEMINI_API_KEY")) if "GEMINI_API_KEY" in st.secrets else None
    if not client: 
        return "KI nicht konfiguriert." if not structured else KiAntwortSchema(wissensluecke_erkannt=True, antwort_text="")
    
    sys_instruction = "Du bist „Villa Avatar“. Antworte kurz, präzise, smartphone-optimiert. Nutze den Excel-Kontext intelligent. Erwähne niemals Tabellenstrukturen."
    try:
        if structured:
            res = client.models.generate_content(model="gemini-2.5-flash", contents=f"Kontext:\n{context}\n\nFrage: {prompt}", config=types.GenerateContentConfig(system_instruction=sys_instruction, temperature=0.2, response_mime_type="application/json", response_schema=KiAntwortSchema))
            data = json.loads(res.text)
            return KiAntwortSchema(wissensluecke_erkannt=bool(data.get("wissensluecke_erkannt", True)), antwort_text=str(data.get("antwort_text", "")))
        
        return client.models.generate_content(model="gemini-2.5-flash", contents=prompt, config=types.GenerateContentConfig(system_instruction="Du bist ein präzises Assistenzsystem für Berichte. Formuliere die Rohdaten sachlich korrekt in 1 bis maximal 2 verständliche Sätze um. Gib NIEMALS Optionen, Auswahlmöglichkeiten, Nummerierungen oder Metatexte aus.", temperature=0.1)).text
    except:
        return KiAntwortSchema(wissensluecke_erkannt=True, antwort_text="") if structured else "Fehler bei Textaufbereitung."

def extract_context_for_object(objekt_name):
    if df_wissen is None or df_lexikon is None or not objekt_name: return ""
    bez_col = df_wissen.columns[0]
    row = df_wissen[df_wissen[bez_col].astype(str).str.strip().str.lower() == objekt_name.lower().strip()]
    if row.empty: return ""
    
    context_parts = [f"Informationen zum Objekt: {objekt_name}"]
    if str(st.session_state.aktive_rolle).lower() == "host":
        for col in df_wissen.columns:
            if col != bez_col and "status" not in col.lower() and col.lower() != "wo?":
                if pd.notna(row.iloc[0][col]) and str(row.iloc[0][col]).strip() != "": 
                    context_parts.append(f"- {col}: {str(row.iloc[0][col]).strip()}")
    else:
        gast_col = next((c for c in df_lexikon.columns if "gast" in c.lower()), df_lexikon.columns[-1])
        tags = df_lexikon[df_lexikon[gast_col].astype(str).str.lower().str.strip() == "ja"][df_lexikon.columns[0]].tolist()
        for col in df_wissen.columns:
            if any(col.lower() == t.lower() for t in tags) and col in row.columns and pd.notna(row.iloc[0][col]) and str(row.iloc[0][col]).strip() != "":
                context_parts.append(f"- {col}: {str(row.iloc[0][col]).strip()}")
                
    st.session_state.last_extracted_context = "\n".join(context_parts)
    return st.session_state.last_extracted_context

# ==============================================================================
# 5. SCHREIBENGINE
# ==============================================================================
def execute_matrix_input_direct(physische_spalte, objekt, text):
    if drive_service is None or df_wissen is None: return
    try:
        fh = io.BytesIO()
        MediaIoBaseDownload(fh, drive_service.files().get_media(fileId=FILE_ID)).next_chunk()
        fh.seek(0)
        wb = openpyxl.load_workbook(fh)
        ws = wb["Wissensbasis"]
        headers = [str(c.value) if c.value else "" for c in ws[1]]
        
        bez_idx = find_column_by_fuzzy_name(headers, "Bezeichnung") or 1
        row_idx = next((r for r in range(2, ws.max_row + 1) if ws.cell(r, bez_idx).value and str(ws.cell(r, bez_idx).value).strip().lower() == objekt.lower().strip()), None)
        if not row_idx:
            row_idx = ws.max_row + 1
            ws.cell(row_idx, bez_idx).value = objekt
            
        col_idx = find_column_by_fuzzy_name(headers, physische_spalte)
        if col_idx:
            old_text = ws.cell(row_idx, col_idx).value or ""
            if old_text:
                ws.cell(row_idx, col_idx).value = f"{str(old_text).strip()}\n{text}".strip()
            else:
                ws.cell(row_idx, col_idx).value = text.strip()
            
            ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True)
            ws.cell(row_idx, col_idx).font = Font(color="1F4E78", bold=False)
            
            status_idx = find_column_by_fuzzy_name(headers, f"{physische_spalte} Status")
            if status_idx:
                old_status = ws.cell(row_idx, status_idx).value or ""
                dat_str = datetime.datetime.now().strftime("%d.%m.%Y")
                neuer_status_eintrag = f"{dat_str} offen"
                if old_status:
                    ws.cell(row_idx, status_idx).value = f"{str(old_status).strip()}\n{neuer_status_eintrag}".strip()
                else:
                    ws.cell(row_idx, status_idx).value = neuer_status_eintrag
            
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            drive_service.files().update(fileId=FILE_ID, media_body=MediaIoBaseUpload(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')).execute()
            st.toast("✅ Matrix aktualisiert!")
    except Exception as e: st.error(f"Schreibfehler: {e}")

def execute_matrix_input(use_case, objekt, text):
    if df_lexikon is None: return
    spalte = next((str(r[df_lexikon.columns[0]]).strip() for _, r in df_lexikon.iterrows() if use_case.lower().strip() in [t.strip().lower() for t in str(r[df_lexikon.columns[4]]).split(',')]), None)
    if not spalte and use_case == "Keine Information": 
        spalte = next((c for c in df_wissen.columns if "information" in c.lower() and "status" not in c.lower()), None)
    if spalte: execute_matrix_input_direct(spalte, objekt, text)

# ==============================================================================
# 6. BENUTZEROBERFLÄCHE (HMI)
# ==============================================================================
st.title("☀️ Villa Avatar")

role = st.selectbox("Rolle", options=["Gast", "Host"], index=None, placeholder="Wer bist du?", label_visibility="collapsed")
if role and role != st.session_state.aktive_rolle:
    st.session_state.aktive_rolle = role
    st.session_state.aktiver_use_case, st.session_state.selected_object, st.session_state.selected_field, st.session_state.messages = None, None, None, []
    st.session_state["erfolgsmeldung_anzeigen"] = None
    st.session_state["host_text_wert"] = ""
    st.session_state.selected_report_type = None
    st.session_state.selected_report_timeframe = None
    st.rerun()

if not st.session_state.aktive_rolle: st.stop()

# Host Authentifizierung
if st.session_state.aktive_rolle == "Host" and not st.session_state.host_authentifiziert:
    pwd = st.text_input("🔑 Passwort eingeben:", type="password")
    if pwd:
        if pwd.strip().lower() == "admin":
            st.session_state.host_authentifiziert = True
            st.session_state.debug_modus_aktiv = True
            st.rerun()
        elif df_passwoerter is not None:
            p_pwd_col = df_passwoerter.columns[1]
            for _, r in df_passwoerter.iterrows():
                if pwd.strip().lower() == str(r[p_pwd_col]).strip().lower():
                    st.session_state.host_authentifiziert = True
                    st.rerun()
    st.stop()

# Menüleiste
if df_usecases is not None:
    uc_col, hmi_col = df_usecases.columns[0], df_usecases.columns[2]
    allowed = [uc for uc in df_usecases[df_usecases[hmi_col].astype(str).str.lower().str.strip() == "ja"][uc_col].tolist() if st.session_state.aktive_rolle == "Host" or any(x in uc.lower() for x in ["hilfe", "störung", "feedback"])]
    cols = st.columns(len(allowed))
    for idx, uc in enumerate(allowed):
        with cols[idx]:
            lbl = next((str(r[df_usecases.columns[3]]).strip() for _, r in df_usecases.iterrows() if str(r[uc_col]).strip() == uc and pd.notna(r[df_usecases.columns[3]])), uc)
            if st.button(lbl, use_container_width=True, type="primary" if st.session_state.aktiver_use_case == uc else "secondary"):
                st.session_state.aktiver_use_case = uc
                st.session_state.selected_object = None
                st.session_state.selected_field = None
                st.session_state.messages = []
                st.session_state["erfolgsmeldung_anzeigen"] = None
                st.session_state["host_text_wert"] = ""
                st.session_state.selected_report_type = None
                st.session_state.selected_report_timeframe = None
                st.rerun()

if not st.session_state.aktiver_use_case: st.stop()

current_uc_clean = str(st.session_state.aktiver_use_case).strip().lower()

# ==============================================================================
# 🎯 USE CASE: HOST MANUAL INPUT
# ==============================================================================
if "information" in current_uc_clean and "keine" not in current_uc_clean and "bericht" not in current_uc_clean and str(st.session_state.aktive_rolle).strip().lower() == "host":
    bez_col, kat_col = df_wissen.columns[0], ("Wo?" if "Wo?" in df_wissen.columns else df_wissen.columns[1])
    def get_liste_host(pattern):
        mask = df_wissen[kat_col].astype(str).str.contains(pattern, case=False, na=False)
        return sorted(df_wissen[mask][bez_col].dropna().drop_duplicates().astype(str).str.strip().tolist())

    tab_innen, tab_aussen, tab_naehe = st.tabs(["🏠 Ausstattung innen", "🌳 Ausstattung außen", "📍 In der Nähe"])
    current_obj = st.session_state.selected_object
    
    with tab_innen:
        options_innen = get_liste_host("innen")
        idx_innen = options_innen.index(current_obj) if current_obj in options_innen else None
        val_innen = st.selectbox("Ausstattung innen", options=options_innen, index=idx_innen, placeholder="Bitte wähle das Objekt aus", key="h_innen", label_visibility="collapsed")
    with tab_aussen:
        options_aussen = get_liste_host("außen|aussen")
        idx_aussen = options_aussen.index(current_obj) if current_obj in options_aussen else None
        val_aussen = st.selectbox("Ausstattung außen", options=options_aussen, index=idx_aussen, placeholder="Bitte wähle das Objekt aus", key="h_aussen", label_visibility="collapsed")
    with tab_naehe:
        options_naehe = get_liste_host("nähe|naehe")
        idx_naehe = options_naehe.index(current_obj) if current_obj in options_naehe else None
        val_naehe = st.selectbox("In der Nähe", options=options_naehe, index=idx_naehe, placeholder="Bitte wähle das Objekt aus", key="h_naehe", label_visibility="collapsed")

    if val_innen and val_innen != st.session_state.selected_object:
        st.session_state.selected_object = val_innen; st.session_state.selected_field = None; st.session_state["erfolgsmeldung_anzeigen"] = None; st.rerun()
    elif val_aussen and val_aussen != st.session_state.selected_object:
        st.session_state.selected_object = val_aussen; st.session_state.selected_field = None; st.session_state["erfolgsmeldung_anzeigen"] = None; st.rerun()
    elif val_naehe and val_naehe != st.session_state.selected_object:
        st.session_state.selected_object = val_naehe; st.session_state.selected_field = None; st.session_state["erfolgsmeldung_anzeigen"] = None; st.rerun()

    if st.session_state.selected_object:
        if df_lexikon is not None:
            lexikon_spalten = df_lexikon[df_lexikon.columns[0]].dropna().astype(str).str.strip().tolist()
            options_spalten = [col for col in lexikon_spalten if col.lower() not in ["spaltenname", "bezeichnung", "wo?", "relevanz gast", "system", "objekt"] and not col.lower().endswith("status")]
        else:
            options_spalten = [c for c in df_wissen.columns if c.lower() not in ["bezeichnung", "wo?", "id", "kategorie", "relevanz gast", "system"] and not c.lower().endswith("status")]

        default_idx = options_spalten.index(st.session_state.selected_field) if st.session_state.selected_field in options_spalten else None
        s_auswahl = st.selectbox("Art der Information", options=options_spalten, index=default_idx, placeholder="Bitte wähle die Art der Information aus", key="dropdown_neue_info_spalte_clean", label_visibility="collapsed")
        
        if s_auswahl and s_auswahl != st.session_state.selected_field:
            st.session_state.selected_field = s_auswahl; st.session_state["erfolgsmeldung_anzeigen"] = None; st.rerun()
        
        if st.session_state.selected_field:
            txt = st.text_area("Inhalt erfassen", value=st.session_state["host_text_wert"], placeholder="Hier den Text eingeben...", label_visibility="collapsed", key="host_text_eingabe")
            if txt != st.session_state["host_text_wert"]: st.session_state["host_text_wert"] = txt

            if st.button("💾 In Excel-Zentralmatrix speichern", type="primary") and txt.strip():
                execute_matrix_input_direct(st.session_state.selected_field, st.session_state.selected_object, txt.strip())
                danke_text = "Vielen Dank für deine Information."
                if df_usecases is not None:
                    uc_row = df_usecases[df_usecases[df_usecases.columns[0]].astype(str).str.lower().str.strip() == current_uc_clean]
                    if not uc_row.empty and len(df_usecases.columns) > 5 and pd.notna(uc_row.iloc[0][df_usecases.columns[5]]):
                        danke_text = str(uc_row.iloc[0][df_usecases.columns[5]]).strip()
                st.session_state["erfolgsmeldung_anzeigen"] = danke_text
                st.session_state["host_text_wert"] = ""
                st.rerun()

            if "erfolgsmeldung_anzeigen" in st.session_state and st.session_state["erfolgsmeldung_anzeigen"]:
                st.success(st.session_state["erfolgsmeldung_anzeigen"])
    st.stop()

# ==============================================================================
# 📊 USE CASE: BERICHTSENGINE
# ==============================================================================
elif "bericht" in current_uc_clean:
    report_options = []
    mapping_dropdown_zu_lexikon_zeile = {}

    if df_lexikon is not None:
        col_spaltenname, col_usecase, col_details = df_lexikon.columns[0], df_lexikon.columns[3], df_lexikon.columns[5]
        df_bericht_rows = df_lexikon[df_lexikon[col_usecase].astype(str).str.strip().str.lower() == "bericht"]

        for _, row_lex in df_bericht_rows.iterrows():
            raw_detail = str(row_lex[col_details]).strip()
            if raw_detail and raw_detail.lower() != "nan":
                clean_opt = raw_detail.split("(")[0].strip()
                if clean_opt and clean_opt not in report_options:
                    report_options.append(clean_opt)
                    
                    mapping_dropdown_zu_lexikon_zeile[clean_opt] = {
                        "spalte_wissen": str(row_lex[col_spaltenname]).strip(),
                        "such_zustand": "offen" if "offen" in raw_detail.lower() else "ok"
                    }

    if not report_options:
        report_options = ["Offene Störungen", "Behobene Störungen", "Offene Wartungen", "Erfolgte Wartungen", "Offenes Feedback", "Behobenes Feedback", "Offene Wissenslücken"]

    idx_report = report_options.index(st.session_state.selected_report_type) if st.session_state.selected_report_type in report_options else None
    selected_rep = st.selectbox("Berichtsart", options=report_options, index=idx_report, placeholder="Bitte wähle die Art des Berichtes.", label_visibility="collapsed", key="report_type_dropdown")
    
    if selected_rep != st.session_state.selected_report_type:
        st.session_state.selected_report_type = selected_rep; st.rerun()

    if st.session_state.selected_report_type:
        timeframe_options = ["1 Woche", "1 Monat", "3 Monate", "1 Jahr"]
        idx_timeframe = timeframe_options.index(st.session_state.selected_report_timeframe) if st.session_state.selected_report_timeframe in timeframe_options else None
        selected_tf = st.selectbox("Zeitraum", options=timeframe_options, index=idx_timeframe, placeholder="Bitte wähle den Zeitraum.", label_visibility="collapsed", key="report_timeframe_dropdown")
        
        if selected_tf != st.session_state.selected_report_timeframe:
            st.session_state.selected_report_timeframe = selected_tf; st.rerun()

        if st.session_state.selected_report_timeframe:
            st.markdown(f"### 📋 {st.session_state.selected_report_type} ({st.session_state.selected_report_timeframe})")

            heute = datetime.datetime.now()
            delta_days = {"1 Woche": 7, "1 Monat": 30, "3 Monate": 90, "1 Jahr": 365}.get(st.session_state.selected_report_timeframe, 30)
            stichtag = heute - datetime.timedelta(days=delta_days)
            is_winterpause = heute.month in [11, 12, 1, 2, 3]

            report_rows = []
            bez_col = df_wissen.columns[0]
            lexikon_meta = mapping_dropdown_zu_lexikon_zeile.get(st.session_state.selected_report_type)
            
            ziel_spalte, target_keyword, standard_fallback_periode = None, "offen", 180
            if lexikon_meta:
                ziel_spalte = lexikon_meta["spalte_wissen"]
                target_keyword = lexikon_meta["such_zustand"]
            else:
                rep_lower = st.session_state.selected_report_type.lower()
                if "störung" in rep_lower: ziel_spalte = "Störung"
                elif "wartung" in rep_lower: ziel_spalte = "Wartung"
                elif "feedback" in rep_lower: ziel_spalte = "Feedback"
                elif "wissenslücke" in rep_lower: ziel_spalte = "Keine Information"
                if "behoben" in rep_lower or "erfolgt" in rep_lower: target_keyword = "ok"

            echte_ziel_spalte = next((c for c in df_wissen.columns if str(c).strip().lower() == str(ziel_spalte).strip().lower()), None)
            echte_status_spalte = next((c for c in df_wissen.columns if str(c).strip().lower() == f"{str(echte_ziel_spalte).strip().lower()} status"), None)

            if st.session_state.debug_modus_aktiv:
                st.info("⚙ **Debug-Modus Aktiv**")
                st.write(f"Zieldatei-Spalten-Mapping: `{ziel_spalte}` -> `{echte_ziel_spalte}` | Status -> `{echte_status_spalte}`")

            if echte_ziel_spalte and echte_status_spalte:
                for idx, row in df_wissen.iterrows():
                    status_val = str(row[echte_status_spalte]).strip() if pd.notna(row[echte_status_spalte]) else ""
                    if not status_val or status_val.lower() == "nan": continue
                    
                    events = parse_status_history(status_val)
                    if not events: continue
                    
                    letztes_event = events[-1]
                    haupt_text = str(row[echte_ziel_spalte]).strip() if pd.notna(row[echte_ziel_spalte]) else ""
                    if not haupt_text or haupt_text.lower() == "nan": continue

                    ist_wartungs_report = "wartung" in echte_ziel_spalte.lower()
                    
                    if ist_wartungs_report:
                        effektive_tage = get_effective_days_excluding_winter(letztes_event["datum"], heute)
                        
                        # 1. Frequenz-Text DIREKT aus der Wissensbasis-Zeile des aktuellen Objekts auslesen
                        vorgabe_text = ""
                        col_details_wartung_wissen = next((c for c in df_wissen.columns if "details zur wartung" in c.lower()), None)
                        
                        if col_details_wartung_wissen and pd.notna(row[col_details_wartung_wissen]):
                            vorgabe_text = str(row[col_details_wartung_wissen]).strip()

                        # 2. Periode aus dem gefundenen Gerätetext ("wöchentlich...") ermitteln
                        aktive_periode = parse_period_from_text(vorgabe_text)
                        if aktive_periode is None:
                            aktive_periode = standard_fallback_periode

                        if st.session_state.debug_modus_aktiv:
                            st.write(f"🔍 **Wartung Check** [{row[bez_col]}]: Letzter Eintrag: {letztes_event['datum'].strftime('%d.%m.%Y')} ({letztes_event['zustand']}) | Effektive Tage: {effektive_tage} | Text aus Wissensbasis: '{vorgabe_text}' | Erkannte Periode: {aktive_periode} Tage | Winterpause aktuell: {is_winterpause}")

                        if target_keyword == "offen":
                            if effektive_tage > aktive_periode and not is_winterpause:
                                aufbereiteter_text = call_gemini(
                                    prompt=f"Formuliere einen kurzen Berichtssatz, dass die Wartung für '{row[bez_col]}' fällig ist (Letzte Wartung war vor {effektive_tage} echten Tagen am {letztes_event['datum'].strftime('%d.%m.%Y')}, Limit ist {aktive_periode} Tage). Keine Optionen.",
                                    structured=False
                                )
                                report_rows.append({"Eintrag": aufbereiteter_text})
                        else:
                            if letztes_event["zustand"] == "ok" and letztes_event["datum"] >= stichtag:
                                aufbereiteter_text = call_gemini(
                                    prompt=f"Formuliere sachlich, dass die Wartung für '{row[bez_col]}' am {letztes_event['datum'].strftime('%d.%m.%Y')} erfolgreich durchgeführt wurde. Text: {haupt_text}",
                                    structured=False
                                )
                                report_rows.append({"Eintrag": aufbereiteter_text})
                    else:
                        if st.session_state.debug_modus_aktiv:
                            st.write(f"🔍 **Standard Check** [{row[bez_col]}]: Letzter Zustand: {letztes_event['zustand']} (Gesucht: {target_keyword}) | Datum: {letztes_event['datum'].strftime('%d.%m.%Y')} (Stichtag: {stichtag.strftime('%d.%m.%Y')})")

                        if letztes_event["zustand"] == target_keyword.lower() and letztes_event["datum"] >= stichtag:
                            aufbereiteter_text = call_gemini(
                                prompt=f"Bringe folgende Information für das Objekt '{row[bez_col]}' in 1 bis maximal 2 verständliche, rein sachliche Berichtssätze. Keine Optionen, Metatexte oder Aufzählungen. Text: {haupt_text}",
                                structured=False
                            )
                            report_rows.append({"Eintrag": aufbereiteter_text})

            if report_rows:
                st.markdown("---")
                st.markdown(f"**Diese {st.session_state.selected_report_type.lower()} sind gemeldet:**")
                for row_data in report_rows: 
                    st.markdown(f"- {row_data['Eintrag']}")
            else:
                st.info("Keine Einträge für diesen Zeitraum gefunden.")
    st.stop()

# ==============================================================================
# 7. CHAT & INPUT INTERFACE
# ==============================================================================
else:
    uc_row = df_usecases[df_usecases[df_usecases.columns[0]].astype(str).str.lower().str.strip() == current_uc_clean]
    direction = str(uc_row.iloc[0][df_usecases.columns[1]]).strip().upper() if not uc_row.empty else "OUTPUT"
    fragetext = str(uc_row.iloc[0][df_usecases.columns[4]]).strip() if not uc_row.empty and len(df_usecases.columns) > 4 and pd.notna(uc_row.iloc[0][df_usecases.columns[4]]) else "Wie kann ich helfen?"
    danke_tmpl = str(uc_row.iloc[0][df_usecases.columns[5]]).strip() if not uc_row.empty and len(df_usecases.columns) > 5 and pd.notna(uc_row.iloc[0][df_usecases.columns[5]]) else "Danke!"

    bez_col, kat_col = df_wissen.columns[0], ("Wo?" if "Wo?" in df_wissen.columns else df_wissen.columns[1])
    def get_liste(pattern):
        mask = df_wissen[kat_col].astype(str).str.contains(pattern, case=False, na=False)
        if st.session_state.aktive_rolle == "Gast" and "Relevanz Gast" in df_wissen.columns:
            mask = mask & (df_wissen["Relevanz Gast"].astype(str).str.strip().str.lower() == "x")
        return sorted(df_wissen[mask][bez_col].dropna().drop_duplicates().astype(str).str.strip().tolist())

    tab_innen, tab_aussen, tab_naehe = st.tabs(["🏠 Ausstattung innen", "🌳 Ausstattung außen", "📍 In der Nähe"])

    with tab_innen: val_innen = st.selectbox("Ausstattung innen", options=get_liste("innen"), index=None, placeholder="Bitte wähle das Objekt aus.", key="g_innen", label_visibility="collapsed")
    with tab_aussen: val_aussen = st.selectbox("Ausstattung außen", options=get_liste("außen|aussen"), index=None, placeholder="Bitte wähle das Objekt aus.", key="g_aussen", label_visibility="collapsed")
    with tab_naehe: val_naehe = st.selectbox("In der Nähe", options=get_liste("nähe|naehe"), index=None, placeholder="Bitte wähle das Objekt aus.", key="g_naehe", label_visibility="collapsed")

    aktuell_gewaehlt = val_innen or val_aussen or val_naehe

    if aktuell_gewaehlt and aktuell_gewaehlt != st.session_state.selected_object:
        st.session_state.selected_object = aktuell_gewaehlt; st.session_state.messages = []; st.rerun()

    if not st.session_state.selected_object: st.stop()

    for msg in st.session_state.messages:
        with st.chat_message(msg["role"]): st.markdown(msg["content"])
        
    if user_input := st.chat_input(fragetext):
        st.session_state.messages.append({"role": "user", "content": user_input}); st.rerun()
        
    if st.session_state.messages and st.session_state.messages[-1]["role"] == "user":
        u_text = st.session_state.messages[-1]["content"]
        is_not_found = "nicht gefunden" in st.session_state.selected_object.lower()
        
        if direction == "OUTPUT":
            if is_not_found:
                st.session_state.messages.append({"role": "assistant", "content": FALLBACK_SATZ})
                execute_matrix_input("Keine Information", st.session_state.selected_object, u_text)
            else:
                with st.spinner("Antwort wird generiert..."):
                    res = call_gemini(u_text, extract_context_for_object(st.session_state.selected_object))
                    if res.wissensluecke_erkannt or not res.antwort_text or any(p in res.antwort_text.lower() for p in ["keine information", "weiß ich nicht", "leider nein"]):
                        st.session_state.messages.append({"role": "assistant", "content": FALLBACK_SATZ})
                        execute_matrix_input("Keine Information", st.session_state.selected_object, u_text)
                    else:
                        st.session_state.messages.append({"role": "assistant", "content": res.antwort_text})
            st.rerun()
            
        elif direction == "INPUT":
            with st.spinner("Eintrag wird protokoliert..."):
                execute_matrix_input(st.session_state.aktiver_use_case, st.session_state.selected_object, u_text)
                st.session_state.messages.append({"role": "assistant", "content": danke_tmpl.replace("{use_case}", st.session_state.aktiver_use_case)})
                st.rerun()
